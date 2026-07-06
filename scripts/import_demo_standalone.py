#!/usr/bin/env python3
"""独立导入脚本：直接将 Demo 数据写入飞书多维表格，不依赖 scheduler_api 服务。"""
import os, sys, time, json
import openpyxl
import httpx
from datetime import datetime

# ===== 环境变量 =====
APP_ID = os.getenv("FEISHU_APP_ID", "")
APP_SECRET = os.getenv("FEISHU_APP_SECRET", "")
APP_TOKEN = os.getenv("BITABLE_APP_TOKEN", "")
TABLE_MAIN = "tbl06oxGEdMNTEB8"
TABLE_ITEMS = "tblJn5iP6imjzE8h"
TABLE_INV = "tblFZNdEwW50izjh"
BASE = "https://open.feishu.cn/open-apis/bitable/v1"
AUTH_URL = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
BATCH_SIZE = 200

# ===== 获取 token =====
def get_token():
    r = httpx.post(AUTH_URL, json={"app_id": APP_ID, "app_secret": APP_SECRET}, timeout=15)
    d = r.json()
    return d["tenant_access_token"]

# ===== 清空表 =====
def clear_table(token, table_id, table_name):
    """先查所有记录再分批删除"""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    all_ids = []
    page_token = None
    while True:
        params = {"page_size": 500}
        if page_token:
            params["page_token"] = page_token
        r = httpx.get(f"{BASE}/apps/{APP_TOKEN}/tables/{table_id}/records", headers=headers, params=params, timeout=30)
        d = r.json()
        items = d.get("data", {}).get("items", [])
        all_ids.extend([it["record_id"] for it in items])
        if not d.get("data", {}).get("has_more"):
            break
        page_token = d["data"]["page_token"]
    
    if not all_ids:
        print(f"  [{table_name}] 表为空，跳过清理")
        return
    
    print(f"  [{table_name}] 删除 {len(all_ids)} 条旧记录...")
    for i in range(0, len(all_ids), BATCH_SIZE):
        batch = all_ids[i:i+BATCH_SIZE]
        r = httpx.request("DELETE", f"{BASE}/apps/{APP_TOKEN}/tables/{table_id}/records", headers=headers, json={"records": batch}, timeout=30)
        if r.status_code != 200:
            print(f"    删除失败: {r.status_code} {r.text[:200]}")
            break
        print(f"    已删除 {min(i+BATCH_SIZE, len(all_ids))}/{len(all_ids)}")
        time.sleep(0.5)

# ===== 批量写入 =====
def batch_write(token, table_id, table_name, records, field_map=None):
    """写入记录到飞书多维表格"""
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    written = 0
    for i in range(0, len(records), BATCH_SIZE):
        batch = records[i:i+BATCH_SIZE]
        body = {"records": [{"fields": r} for r in batch]}
        r = httpx.post(f"{BASE}/apps/{APP_TOKEN}/tables/{table_id}/records/batch_create", headers=headers, json=body, timeout=60)
        jd = r.json()
        if jd.get("code") != 0:
            print(f"  [{table_name}] 写入失败: code={jd.get('code')} msg={jd.get('msg')}")
            return written
        written += len(batch)
        print(f"  [{table_name}] 已写入 {written}/{len(records)}")
        time.sleep(0.3)
    return written

# ===== 日期格式化 =====
def fmt_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return int(val.timestamp() * 1000)
    try:
        d = datetime.strptime(str(val)[:10], "%Y-%m-%d")
        return int(d.timestamp() * 1000)
    except:
        return None

# ===== 主流程 =====
def main(xlsx_path):
    print(f"读取: {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    
    # 读订单表
    ws1 = wb["订单表"]
    headers1 = [c.value for c in ws1[1]]
    orders = []
    for r in range(2, ws1.max_row + 1):
        vals = [ws1.cell(row=r, column=c+1).value for c in range(len(headers1))]
        orders.append(dict(zip(headers1, vals)))
    print(f"订单表: {len(orders)} 行")
    
    # 读库存表
    ws2 = wb["库存表"]
    headers2 = [c.value for c in ws2[1]]
    inventory = []
    for r in range(2, ws2.max_row + 1):
        vals = [ws2.cell(row=r, column=c+1).value for c in range(len(headers2))]
        inventory.append(dict(zip(headers2, vals)))
    print(f"库存表: {len(inventory)} 行")
    
    # 获取 token
    print("\n获取飞书 Token...")
    token = get_token()
    print("Token OK")
    
    # 1) 清空主表
    print("\n=== 清空旧数据 ===")
    clear_table(token, TABLE_MAIN, "销售订单主表")
    clear_table(token, TABLE_ITEMS, "销售订单明细表")
    clear_table(token, TABLE_INV, "库存快照表")
    
    # 2) 写入主表（仅写入实际存在的字段）
    print("\n=== 写入新数据 ===")
    main_records = []
    for o in orders:
        rec = {
            "合同编号": str(o.get("合同编号", "")),
            "客户名称": str(o.get("客户名称", "")),
            "代理商": str(o.get("代理商", "")),
            "项目名称": str(o.get("所属项目", "")),
            "下单日期": fmt_date(o.get("签订日期")),
            "商务": str(o.get("代理商", "")),  # 复用代理商
        }
        main_records.append(rec)
    written = batch_write(token, TABLE_MAIN, "销售订单主表", main_records)
    
    # 3) 写入明细表
    items_records = []
    for o in orders:
        pname = str(o.get("产品名称", ""))
        if "费" in pname:
            continue
        qty = o.get("数量")
        items_records.append({
            "合同编号": str(o.get("合同编号", "")),
            "产品名称": pname,
            "规格": str(o.get("明细规格", "")),
            "合同数量": float(qty) if qty else 0,
            "下单时间": fmt_date(o.get("签订日期")),
        })
    written = batch_write(token, TABLE_ITEMS, "销售订单明细表", items_records)
    
    # 4) 写入库存表（仅写入实际存在的字段）
    inv_records = []
    now_ts = int(datetime.now().timestamp() * 1000)
    for inv in inventory:
        inv_records.append({
            "国网设备名称": str(inv.get("国网设备名称", "")),
            "国网设备型号": str(inv.get("国网设备型号", "")),
            "库存数量": int(inv.get("库存数量", 0)) if inv.get("库存数量") else 0,
            "待采购出库": int(inv.get("待采购出库", 0)) if inv.get("待采购出库") else 0,
            "在途数量": str(inv.get("锁定总库存", "")),  # 用锁定数作为在途
            "库存日期": now_ts,
        })
    written = batch_write(token, TABLE_INV, "库存快照表", inv_records)
    
    print(f"\n=== 完成 ===")
    print(f"主表: {len(main_records)} 条")
    print(f"明细表: {len(items_records)} 条")
    print(f"库存表: {len(inv_records)} 条")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "/Users/xiaowang/Desktop/测试集_演示.xlsx"
    main(path)
